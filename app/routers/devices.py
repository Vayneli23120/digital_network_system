"""Device management router"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import Optional, List
from pathlib import Path
import shutil
from datetime import datetime
from io import BytesIO

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from starlette.responses import StreamingResponse
except ImportError:
    StreamingResponse = None

from ..database import get_db
from ..models import Device, BackupRecord, FaultRecord, MaintenanceRecord, DevicePhoto, CredentialGroup
from ..config import get_config

config = get_config()

router = APIRouter(prefix="/api/devices", tags=["devices"])


@router.get("")
async def list_devices(status: Optional[str] = None, role: Optional[str] = None):
    """获取设备列表"""
    db: Session = next(get_db())

    try:
        query = db.query(Device)

        if status:
            query = query.filter(Device.status == status)
        if role:
            query = query.filter(Device.role == role)

        devices = query.all()

        return {
            "items": [
                {
                    "id": d.id,
                    "name": d.name,
                    "ip": d.ip,
                    "model": d.model,
                    "serial_number": d.serial_number,
                    "location": d.location,
                    "role": d.role,
                    "status": d.status,
                    "credential_group": d.credential_group,
                    "last_backup_time": d.last_backup_time.isoformat() if d.last_backup_time else None,
                }
                for d in devices
            ]
        }
    finally:
        db.close()


@router.get("/export")
async def export_devices():
    """导出设备信息为 Excel 文件"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel 支持未安装，请运行 pip install openpyxl")

    db: Session = next(get_db())

    try:
        devices = db.query(Device).all()

        # 创建 Excel 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Devices"

        # 表头
        headers = ["name", "ip", "model", "serial_number", "location", "role", "status", "credential_group", "vendor", "purchase_cost"]
        ws.append(headers)

        # 数据
        for device in devices:
            ws.append([
                device.name,
                device.ip or "",
                device.model or "",
                device.serial_number or "",
                device.location or "",
                device.role or "",
                device.status or "",
                device.credential_group or "default",
                device.vendor or "",
                float(device.purchase_cost) if device.purchase_cost else 0
            ])

        # 输出为字节流
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=devices.xlsx"}
        )
    finally:
        db.close()


@router.post("/import")
async def import_devices(file: UploadFile = File(...)):
    """从 Excel 文件导入设备信息"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel 支持未安装，请运行 pip install openpyxl")

    db: Session = next(get_db())

    try:
        # 读取上传的文件
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active

        # 读取表头
        headers = [cell.value for cell in ws[1]]

        success_count = 0
        failed_count = 0
        errors = []

        # 处理数据行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # 跳过空行
                if not any(row):
                    continue

                # 构建设备数据字典
                device_data = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        device_data[header] = row[i]

                # 验证必填字段
                if not device_data.get("name") or not device_data.get("ip"):
                    errors.append(f"第{row_idx}行：缺少必填字段 (name 或 ip)")
                    failed_count += 1
                    continue

                # 检查设备名称是否已存在
                existing = db.query(Device).filter(Device.name == device_data["name"]).first()
                if existing:
                    errors.append(f"第{row_idx}行：设备名称 '{device_data['name']}' 已存在")
                    failed_count += 1
                    continue

                # 创建设备
                device = Device(
                    name=device_data.get("name"),
                    ip=device_data.get("ip"),
                    model=device_data.get("model", ""),
                    serial_number=device_data.get("serial_number", ""),
                    location=device_data.get("location", ""),
                    role=device_data.get("role", "access"),
                    status=device_data.get("status", "online"),
                    credential_group=device_data.get("credential_group", "default"),
                    vendor=device_data.get("vendor", ""),
                    purchase_cost=device_data.get("purchase_cost", 0)
                )
                db.add(device)
                success_count += 1

            except Exception as e:
                errors.append(f"第{row_idx}行：{str(e)}")
                failed_count += 1

        db.commit()

        return {
            "success": success_count,
            "failed": failed_count,
            "errors": errors
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"导入失败：{str(e)}")
    finally:
        db.close()


@router.get("/{device_id}")
async def get_device(device_id: int):
    """获取设备详情"""
    db: Session = next(get_db())

    from loguru import logger

    try:
        device = db.query(Device).filter(Device.id == device_id).first()

        if not device:
            raise HTTPException(status_code=404, detail="设备不存在")

        # 获取设备照片
        photos = db.query(DevicePhoto).filter(DevicePhoto.device_id == device_id).all()

        # 获取最近的备份记录
        backups = db.query(BackupRecord).filter(
            BackupRecord.device_id == device_id
        ).order_by(BackupRecord.backup_time.desc()).limit(5).all()

        # 获取故障记录
        faults = db.query(FaultRecord).filter(
            FaultRecord.device_id == device_id
        ).order_by(FaultRecord.created_at.desc()).limit(5).all()
        logger.info(f"Debug - faults for device {device_id}: {[(f.id, f.status) for f in faults]}")

        # 获取维修记录
        maintenances = db.query(MaintenanceRecord).filter(
            MaintenanceRecord.device_id == device_id
        ).order_by(MaintenanceRecord.created_at.desc()).limit(5).all()

        return {
            "id": device.id,
            "name": device.name,
            "ip": device.ip,
            "model": device.model,
            "serial_number": device.serial_number,
            "location": device.location,
            "role": device.role,
            "status": device.status,
            "purchase_date": device.purchase_date.isoformat() if device.purchase_date else None,
            "vendor": device.vendor,
            "purchase_cost": float(device.purchase_cost) if device.purchase_cost else 0,
            "photos": [
                {
                    "id": p.id,
                    "photo_type": p.photo_type,
                    "photo_path": p.photo_path,
                    "upload_date": p.upload_date.isoformat()
                }
                for p in photos
            ],
            "recent_backups": [
                {
                    "id": b.id,
                    "backup_time": b.backup_time.isoformat(),
                    "has_change": b.has_change
                }
                for b in backups
            ],
            "recent_faults": [
                {
                    "id": f.id,
                    "fault_no": f.fault_no,
                    "device_id": f.device_id,
                    "severity": f.severity,
                    "status": f.status,
                    "description": f.description,
                    "downtime_minutes": f.downtime_minutes,
                    "impact": f.impact,
                    "created_at": f.created_at.isoformat()
                }
                for f in faults
            ],
            "recent_maintenances": [
                {
                    "id": m.id,
                    "maint_no": m.maint_no,
                    "device_id": m.device_id,
                    "maint_type": m.maint_type,
                    "parts_replaced": m.parts_replaced,
                    "parts_cost": m.parts_cost,
                    "labor_hours": m.labor_hours,
                    "labor_cost": m.labor_cost,
                    "vendor": m.vendor,
                    "description": m.description,
                    "maint_time": m.maint_time.isoformat() if m.maint_time else None,
                    "created_at": m.created_at.isoformat()
                }
                for m in maintenances
            ]
        }
    finally:
        db.close()


@router.post("")
async def create_device(device_data: dict):
    """创建新设备"""
    db: Session = next(get_db())

    try:
        device = Device(**device_data)
        db.add(device)
        db.commit()
        db.refresh(device)

        return {"id": device.id, "message": "设备创建成功"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        db.close()


@router.put("/{device_id}")
async def update_device(device_id: int, device_data: dict):
    """更新设备信息"""
    db: Session = next(get_db())

    try:
        device = db.query(Device).filter(Device.id == device_id).first()
        if not device:
            raise HTTPException(status_code=404, detail="设备不存在")

        # 允许更新的字段列表
        allowed_fields = [
            "ip", "model", "serial_number", "location",
            "role", "status", "purchase_date", "vendor", "purchase_cost",
            "photo_dir", "credential_group"
        ]

        # 只更新允许的字段
        for key, value in device_data.items():
            if key in allowed_fields and hasattr(device, key):
                setattr(device, key, value)

        db.commit()
        db.refresh(device)

        return {"id": device.id, "message": "更新成功"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.delete("/{device_id}")
async def delete_device(device_id: int):
    """删除设备"""
    db: Session = next(get_db())

    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    db.delete(device)
    db.commit()

    return {"message": "删除成功"}


# ============ 照片管理 API ============

@router.post("/{device_id}/photos")
async def upload_device_photo(
    device_id: int,
    photo: UploadFile = File(...),
    photo_type: str = "other",
    uploader: str = None
):
    """上传设备照片"""
    db: Session = next(get_db())

    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    # 确保目录存在
    photo_dir = Path(config.storage.photo_dir) / device.name / "photos"
    photo_dir.mkdir(parents=True, exist_ok=True)

    # 保存文件
    filename = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{photo.filename}"
    file_path = photo_dir / filename

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(photo.file, buffer)

    # 记录数据库
    photo_record = DevicePhoto(
        device_id=device_id,
        device_name=device.name,
        photo_path=str(file_path),
        photo_type=photo_type,
        uploader=uploader or "system"
    )
    db.add(photo_record)
    db.commit()
    db.refresh(photo_record)

    return {"id": photo_record.id, "path": str(file_path), "filename": filename}


@router.get("/{device_id}/photos")
async def get_device_photos(device_id: int):
    """获取设备照片列表"""
    db: Session = next(get_db())
    photos = db.query(DevicePhoto).filter(
        DevicePhoto.device_id == device_id
    ).order_by(DevicePhoto.upload_date.desc()).all()

    return {
        "items": [
            {
                "id": p.id,
                "photo_type": p.photo_type,
                "photo_path": p.photo_path,
                "upload_date": p.upload_date.isoformat(),
                "uploader": p.uploader
            }
            for p in photos
        ]
    }


@router.delete("/{device_id}/photos/{photo_id}")
async def delete_device_photo(device_id: int, photo_id: int):
    """删除设备照片"""
    db: Session = next(get_db())
    photo = db.query(DevicePhoto).filter(
        DevicePhoto.id == photo_id,
        DevicePhoto.device_id == device_id
    ).first()

    if not photo:
        raise HTTPException(status_code=404, detail="照片不存在")

    # 删除文件
    photo_path = Path(photo.photo_path)
    if photo_path.exists():
        photo_path.unlink()

    # 删除记录
    db.delete(photo)
    db.commit()

    return {"message": "删除成功"}
