"""Device management router"""

import asyncio
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import Optional, List, Literal
from pathlib import Path
import shutil
from datetime import datetime
from io import BytesIO

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from starlette.responses import StreamingResponse
except ImportError:
    StreamingResponse = None

from pydantic import BaseModel

from app.shared.database import get_db
from app.shared.models import Device, BackupRecord, FaultRecord, MaintenanceRecord, DevicePhoto, CredentialGroup, SparePartInstance, SparePart, DeviceInterface, InterfaceTrafficSample
from app.shared.config import get_config

config = get_config()

router = APIRouter(prefix="/api/devices", tags=["devices"])


# ============ Pydantic 模型 ============

class DeviceCreate(BaseModel):
    name: str
    ip: str
    model: Optional[str] = None
    location: Optional[str] = None
    device_type: str = "other"
    role: str = "access"
    # 新字段：部署状态
    deployment_status: str = "un-used"  # in-use, un-used, maintenance, retired
    # 兼容旧字段（将被映射到 deployment_status）
    status: Optional[str] = None
    vendor: Optional[str] = None
    purchase_cost: float = 0
    credential_group: str = "default"
    # 监控分级：critical(10s/3周期) / normal(30s/2周期) / low(120s/2周期)
    monitor_tier: Literal["critical", "normal", "low"] = "normal"
    modules: Optional[List[dict]] = None  # [{"type": "main", "serial_number": "SN001"}]


class DeviceUpdate(BaseModel):
    ip: Optional[str] = None
    model: Optional[str] = None
    location: Optional[str] = None
    device_type: Optional[str] = None
    role: Optional[str] = None
    # 新字段：部署状态
    deployment_status: Optional[str] = None
    # 兼容旧字段（将被映射到 deployment_status）
    status: Optional[str] = None
    vendor: Optional[str] = None
    purchase_cost: Optional[float] = None
    credential_group: Optional[str] = None
    name: Optional[str] = None
    monitor_tier: Optional[Literal["critical", "normal", "low"]] = None
    modules: Optional[List[dict]] = None


class ProbeRequest(BaseModel):
    """设备探测请求"""
    ip: str
    credential_group: Optional[str] = None
    vendor: Optional[str] = "cisco"
    device_type: Optional[str] = None


# ============ 设备预检 API ============

@router.post("/test-reachability")
async def test_device_reachability(request: ProbeRequest):
    """测试设备IP可达性（ping测试）

    所有设备类型都可用此API测试网络连通性。
    """
    from .device_probe_service import get_probe_service
    service = get_probe_service()
    return service.test_ip_reachability(request.ip)


@router.post("/test-connection")
async def test_device_connection(request: ProbeRequest):
    """测试设备SSH连接

    仅支持SSH的设备类型可用此API。
    AP设备不支持SSH，防火墙需要特殊权限。
    """
    from .device_probe_service import get_probe_service
    db = next(get_db())
    try:
        service = get_probe_service()
        return service.test_ssh_connection(
            db,
            request.ip,
            request.credential_group or "default",
            request.vendor or "cisco",
            request.device_type
        )
    finally:
        db.close()


@router.post("/fetch-info")
async def fetch_device_info(request: ProbeRequest):
    """获取设备信息

    通过SSH连接设备执行show inventory和show snmp获取信息。
    仅支持SSH的设备类型可用此API。
    """
    from .device_probe_service import get_probe_service
    db = next(get_db())
    try:
        service = get_probe_service()
        return service.fetch_device_info(
            db,
            request.ip,
            request.credential_group or "default",
            request.vendor or "cisco",
            request.device_type
        )
    finally:
        db.close()


@router.get("")
async def list_devices(status: Optional[str] = None, role: Optional[str] = None,
                        device_type: Optional[str] = None,
                        deployment_status: Optional[str] = None,
                        reachability: Optional[str] = None,
                        skip: int = 0, limit: int = 200,
                        db: Session = Depends(get_db)):
    """获取设备列表

    Args:
        status: 按旧状态过滤（兼容）
        deployment_status: 按部署状态过滤 (in-use/un-used/maintenance/retired)
        reachability: 按可达性过滤 (reachable/unreachable/unknown)
    """
    from .device_service import list_devices as svc_list_devices
    return svc_list_devices(db, status=status, role=role, device_type=device_type,
                           deployment_status=deployment_status, reachability=reachability,
                           skip=skip, limit=limit)


@router.get("/export")
async def export_devices():
    """导出设备信息为 Excel 文件"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel 支持未安装，请运行 pip install openpyxl")

    db: Session = next(get_db())

    try:
        devices = db.query(Device).order_by(Device.id).limit(5000).all()

        # 创建 Excel 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Devices"

        # 表头
        headers = ["name", "ip", "model", "serial_number", "location", "device_type", "role",
                   "deployment_status", "reachability", "credential_group", "vendor", "purchase_cost"]
        ws.append(headers)

        # 数据
        for device in devices:
            ws.append([
                device.name,
                device.ip or "",
                device.model or "",
                device.serial_number or "",
                device.location or "",
                device.device_type or "other",
                device.role or "",
                device.deployment_status or "un-used",
                device.reachability or "unknown",
                device.credential_group or "default",
                device.vendor or "",
                float(device.purchase_cost) if device.purchase_cost else 0
            ])

        # 输出为字节流
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=devices.xlsx"}
        )
    finally:
        db.close()


@router.post("/import")
async def import_devices(file: UploadFile = File(...)):
    """从 Excel 文件导入设备信息"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel 支持未安装，请运行 pip install openpyxl")

    db: Session = next(get_db())

    try:
        # 读取上传的文件
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active

        # 读取表头
        headers = [cell.value for cell in ws[1]]

        success_count = 0
        failed_count = 0
        errors = []

        # 处理数据行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # 跳过空行
                if not any(row):
                    continue

                # 构建设备数据字典
                device_data = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        device_data[header] = row[i]

                # 验证必填字段
                if not device_data.get("name") or not device_data.get("ip"):
                    errors.append(f"第{row_idx}行：缺少必填字段 (name 或 ip)")
                    failed_count += 1
                    continue

                # 检查设备名称是否已存在
                existing = db.query(Device).filter(Device.name == device_data["name"]).first()
                if existing:
                    errors.append(f"第{row_idx}行：设备名称 '{device_data['name']}' 已存在")
                    failed_count += 1
                    continue

                # 创建设备
                device = Device(
                    name=device_data.get("name"),
                    ip=device_data.get("ip"),
                    model=device_data.get("model", ""),
                    serial_number=device_data.get("serial_number", ""),
                    location=device_data.get("location", ""),
                    role=device_data.get("role", "access"),
                    # 新字段
                    deployment_status=device_data.get("deployment_status", "un-used"),
                    reachability=device_data.get("reachability", "unknown"),
                    # 兼容旧字段
                    status=device_data.get("status", "online"),
                    credential_group=device_data.get("credential_group", "default"),
                    vendor=device_data.get("vendor", ""),
                    device_type=device_data.get("device_type", "other"),
                    purchase_cost=device_data.get("purchase_cost", 0)
                )
                db.add(device)
                success_count += 1

            except Exception as e:
                errors.append(f"第{row_idx}行：{str(e)}")
                failed_count += 1

        db.commit()

        return {
            "success": success_count,
            "failed": failed_count,
            "errors": errors
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"导入失败：{str(e)}")
    finally:
        db.close()


# ============ 厂商管理 API（放在 /{device_id} 之前，避免路由匹配冲突）============

@router.get("/vendors")
async def list_vendors():
    """获取支持的厂商列表"""
    from .vendor_service import get_supported_vendors
    return get_supported_vendors()


@router.get("/vendors/{vendor}")
async def get_vendor(vendor: str):
    """获取厂商详细信息"""
    from .vendor_service import get_vendor_info
    return get_vendor_info(vendor)


# ============ 可达性监控 API ============

@router.post("/{device_id}/check-reachability")
async def manual_check_reachability(device_id: int):
    """手动触发单设备可达性检测

    立即检测设备可达性并返回结果。
    """
    from app.services.reachability_monitor import get_reachability_monitor

    db = next(get_db())
    try:
        device = db.query(Device).filter(Device.id == device_id).first()
        if not device:
            raise HTTPException(status_code=404, detail="设备不存在")

        monitor = get_reachability_monitor()
        result = monitor.check_device_reachability(db, device)

        return {
            "device_id": device_id,
            "device_name": device.name,
            "reachability": device.reachability,
            "latency_ms": device.reachability_latency_ms,
            "method": device.reachability_method,
            "last_check": device.last_reachability_check.isoformat() if device.last_reachability_check else None,
        }
    finally:
        db.close()


@router.get("/reachability-stats")
async def get_reachability_stats():
    """获取可达性统计

    返回已部署设备的可达性状态统计。
    """
    db = next(get_db())
    try:
        total = db.query(Device).filter(Device.deployment_status == 'in-use').count()
        reachable = db.query(Device).filter(
            Device.deployment_status == 'in-use',
            Device.reachability == 'reachable'
        ).count()
        unreachable = db.query(Device).filter(
            Device.deployment_status == 'in-use',
            Device.reachability == 'unreachable'
        ).count()
        unknown = db.query(Device).filter(
            Device.deployment_status == 'in-use',
            Device.reachability == 'unknown'
        ).count()

        # 按设备类型统计
        device_types = ['uce', 'core_switch', 'server_switch', 'office_switch', 'ap', 'wlc', 'router', 'pa', 'ftd', 'other']
        by_type = {}
        for dtype in device_types:
            type_total = db.query(Device).filter(
                Device.device_type == dtype,
                Device.deployment_status == 'in-use'
            ).count()
            by_type[dtype] = {
                'total': type_total,
                'reachable': db.query(Device).filter(
                    Device.device_type == dtype,
                    Device.deployment_status == 'in-use',
                    Device.reachability == 'reachable'
                ).count(),
                'unreachable': db.query(Device).filter(
                    Device.device_type == dtype,
                    Device.deployment_status == 'in-use',
                    Device.reachability == 'unreachable'
                ).count(),
            }

        return {
            "total_deployed": total,
            "reachable": reachable,
            "unreachable": unreachable,
            "unknown": unknown,
            "online_rate": round(reachable / total * 100, 2) if total > 0 else 0,
            "by_type": by_type,
        }
    finally:
        db.close()


@router.get("/monitor/status")
async def get_monitor_status():
    """获取可达性监控服务状态"""
    from app.services.reachability_monitor import get_reachability_monitor
    monitor = get_reachability_monitor()
    return monitor.get_stats()


@router.get("/monitor/diagnostics")
async def get_monitor_diagnostics():
    """可达性监控诊断

    列出哪些设备正在被监控、哪些被跳过（及原因），以及每台设备的当前状态与检测历史。
    用于排查"设备离线但大屏无反应"的问题。
    """
    from app.services.reachability_monitor import get_reachability_monitor
    monitor = get_reachability_monitor()
    return monitor.diagnostics()


@router.get("/monitor/trap-diagnostics")
async def get_trap_diagnostics():
    """SNMP Trap 接收器诊断（是否在监听、收到/应用的 trap 数量）"""
    from app.services.trap_receiver import get_trap_receiver
    return get_trap_receiver().diagnostics()


@router.post("/monitor/check-now")
async def trigger_monitor_check_now(tier: Optional[str] = None):
    """立即触发一次可达性检测（测试用，免等定时周期）

    Args:
        tier: 可选，指定分级 critical/normal/low，不传则检测全部分级
    """
    import asyncio as _asyncio
    from app.services.reachability_monitor import get_reachability_monitor
    monitor = get_reachability_monitor()
    # 在工作线程中运行（check_now 内部使用 asyncio.run，不能在运行中的事件循环里调用）
    summary = await _asyncio.to_thread(monitor.check_now, tier)
    return {"triggered": True, "tier": tier or "all", "result": summary}


@router.get("/{device_id}/metrics")
async def get_device_performance_metrics(device_id: int, db: Session = Depends(get_db)):
    """获取设备性能指标

    通过 SNMP 查询获取 CPU、内存、温度、上行链路带宽利用率。

    Returns:
        {
            "cpu": {"value": 45, "status": "normal"},
            "memory": {"used_percent": 60, "used_mb": 512, "total_mb": 1024, "status": "normal"},
            "temperature": {"value": 35, "status": "normal", "threshold": 70},
            "uplinks": [{"interface": "Gi1/0/1", "alias": "上行核心", "utilization": 25, ...}],
            "timestamp": "2024-01-15T10:30:00",
            "snmp_available": true
        }
    """
    import asyncio
    from .snmp_service import get_snmp_service

    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    if not device.ip:
        raise HTTPException(status_code=400, detail="设备未配置 IP 地址")

    community = device.snmp_community or "public"
    vendor = device.vendor or "cisco"

    # 未启用 SNMP 或未配置社区串时直接返回，避免无谓的超时等待
    if not device.snmp_enabled or not device.snmp_community:
        return {
            "cpu": {"value": None, "status": "unknown", "message": "未启用 SNMP 或未配置社区串"},
            "memory": {"value": None, "status": "unknown"},
            "temperature": {"value": None, "status": "unknown"},
            "interfaces": {"up": None, "down": None, "total": None},
            "errors": {"total_errors": None, "has_errors": False},
            "uplinks": [],
            "uptime": {"uptime_days": None, "human": None},
            "snmp_available": False,
            "device_ip": device.ip,
            "error": "设备未启用 SNMP 或未配置社区串，请在设备配置中填写"
        }

    service = get_snmp_service()

    if not service.is_available():
        return {
            "cpu": {"value": None, "status": "unknown", "message": "SNMP 服务未安装"},
            "memory": {"value": None, "status": "unknown"},
            "temperature": {"value": None, "status": "unknown"},
            "interfaces": {"up": None, "down": None, "total": None},
            "errors": {"total_errors": None, "has_errors": False},
            "uplinks": [],
            "uptime": {"uptime_days": None, "human": None},
            "snmp_available": False,
            "device_ip": device.ip
        }

    try:
        # 直接调用异步方法
        metrics = await asyncio.wait_for(
            service.get_device_metrics_async(device.ip, community, vendor),
            timeout=12.0
        )
        metrics["snmp_available"] = True
        metrics["device_ip"] = device.ip
        return metrics
    except asyncio.TimeoutError:
        logger.warning(f"SNMP query timeout for device {device.ip}")
        return {
            "cpu": {"value": None, "status": "unknown", "message": "查询超时"},
            "memory": {"value": None, "status": "unknown"},
            "temperature": {"value": None, "status": "unknown"},
            "interfaces": {"up": None, "down": None, "total": None},
            "errors": {"total_errors": None, "has_errors": False},
            "uplinks": [],
            "uptime": {"uptime_days": None, "human": None},
            "snmp_available": True,
            "device_ip": device.ip,
            "error": "SNMP 查询超时，设备可能未配置 SNMP 或网络不可达"
        }
    except Exception as e:
        logger.error(f"SNMP query failed for device {device.ip}: {e}")
        return {
            "cpu": {"value": None, "status": "unknown", "message": str(e)},
            "memory": {"value": None, "status": "unknown"},
            "temperature": {"value": None, "status": "unknown"},
            "interfaces": {"up": None, "down": None, "total": None},
            "errors": {"total_errors": None, "has_errors": False},
            "uplinks": [],
            "uptime": {"uptime_days": None, "human": None},
            "snmp_available": True,
            "device_ip": device.ip,
            "error": str(e)
        }


@router.get("/{device_id}")
async def get_device(device_id: int, db: Session = Depends(get_db)):
    """获取设备详情"""
    from .device_service import get_device as svc_get_device
    from app.shared.exceptions import ResourceNotFoundException
    try:
        device = svc_get_device(db, device_id)
        photos = db.query(DevicePhoto).filter(DevicePhoto.device_id == device_id).all()
        backups = db.query(BackupRecord).filter(BackupRecord.device_id == device_id).order_by(BackupRecord.backup_time.desc()).limit(5).all()
        faults = db.query(FaultRecord).filter(FaultRecord.device_id == device_id).order_by(FaultRecord.created_at.desc()).limit(5).all()
        maintenances = db.query(MaintenanceRecord).filter(MaintenanceRecord.device_id == device_id).order_by(MaintenanceRecord.created_at.desc()).limit(5).all()

        return {
            **device,
            "photos": [{"id": p.id, "photo_type": p.photo_type, "photo_path": p.photo_path, "upload_date": p.upload_date.isoformat()} for p in photos],
            "recent_backups": [{"id": b.id, "backup_time": b.backup_time.isoformat(), "has_change": b.has_change} for b in backups],
            "recent_faults": [{"id": f.id, "fault_no": f.fault_no, "severity": f.severity, "status": f.status, "description": f.description, "downtime_minutes": f.downtime_minutes, "impact": f.impact, "created_at": f.created_at.isoformat()} for f in faults],
            "recent_maintenances": [{"id": m.id, "maint_no": m.maint_no, "maint_type": m.maint_type, "parts_replaced": m.parts_replaced, "parts_cost": float(m.parts_cost or 0), "labor_hours": float(m.labor_hours or 0), "labor_cost": float(m.labor_cost or 0), "vendor": m.vendor, "description": m.description, "maint_time": m.maint_time.isoformat() if m.maint_time else None, "created_at": m.created_at.isoformat()} for m in maintenances],
        }
    except ResourceNotFoundException:
        raise HTTPException(status_code=404, detail="设备不存在")


@router.post("")
async def create_device(device_data: DeviceCreate, db: Session = Depends(get_db)):
    """创建新设备"""
    from .device_service import create_device as svc_create_device
    from app.shared.exceptions import ConflictException
    try:
        result = svc_create_device(db, device_data.model_dump())
        from app.shared.cache import cache
        cache.invalidate_prefix("dashboard:")
        return {"id": result["id"], "message": "设备创建成功"}
    except ConflictException as e:
        raise HTTPException(status_code=409, detail=str(e))


@router.put("/{device_id}")
async def update_device(device_id: int, device_data: DeviceUpdate, db: Session = Depends(get_db)):
    """更新设备信息"""
    from .device_service import update_device as svc_update_device
    from app.shared.exceptions import ResourceNotFoundException
    try:
        result = svc_update_device(db, device_id, device_data.model_dump(exclude_unset=True))
        return {"id": result["id"], "message": "更新成功"}
    except ResourceNotFoundException:
        raise HTTPException(status_code=404, detail="设备不存在")


@router.delete("/{device_id}")
async def delete_device(device_id: int, db: Session = Depends(get_db)):
    """删除设备"""
    from .device_service import delete_device as svc_delete_device
    from app.shared.exceptions import ResourceNotFoundException
    try:
        result = svc_delete_device(db, device_id)
        from app.shared.cache import cache
        cache.invalidate_prefix("dashboard:")
        return result
    except ResourceNotFoundException:
        raise HTTPException(status_code=404, detail="设备不存在")


# ============ 照片管理 API ============

@router.post("/{device_id}/photos")
async def upload_device_photo(
    device_id: int,
    photo: UploadFile = File(...),
    photo_type: str = "other",
    uploader: str = None
):
    """上传设备照片"""
    db: Session = next(get_db())

    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    # 确保目录存在
    photo_dir = Path(config.storage.photo_dir) / device.name / "photos"
    photo_dir.mkdir(parents=True, exist_ok=True)

    # 保存文件
    filename = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{photo.filename}"
    file_path = photo_dir / filename

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(photo.file, buffer)

    # 记录数据库
    photo_record = DevicePhoto(
        device_id=device_id,
        device_name=device.name,
        photo_path=str(file_path),
        photo_type=photo_type,
        uploader=uploader or "system"
    )
    db.add(photo_record)
    db.commit()
    db.refresh(photo_record)

    return {"id": photo_record.id, "path": str(file_path), "filename": filename}


@router.get("/{device_id}/photos")
async def get_device_photos(device_id: int):
    """获取设备照片列表"""
    db: Session = next(get_db())
    photos = db.query(DevicePhoto).filter(
        DevicePhoto.device_id == device_id
    ).order_by(DevicePhoto.upload_date.desc()).all()

    return {
        "items": [
            {
                "id": p.id,
                "photo_type": p.photo_type,
                "photo_path": p.photo_path,
                "upload_date": p.upload_date.isoformat(),
                "uploader": p.uploader
            }
            for p in photos
        ]
    }


@router.delete("/{device_id}/photos/{photo_id}")
async def delete_device_photo(device_id: int, photo_id: int):
    """删除设备照片"""
    db: Session = next(get_db())
    photo = db.query(DevicePhoto).filter(
        DevicePhoto.id == photo_id,
        DevicePhoto.device_id == device_id
    ).first()

    if not photo:
        raise HTTPException(status_code=404, detail="照片不存在")

    # 删除文件
    photo_path = Path(photo.photo_path)
    if photo_path.exists():
        photo_path.unlink()

    # 删除记录
    db.delete(photo)
    db.commit()

    return {"message": "删除成功"}


@router.get("/{device_id}/inventory")
async def get_device_inventory(device_id: int, db: Session = Depends(get_db)):
    """获取设备当前安装的备件列表

    Args:
        device_id: 设备 ID

    Returns:
        设备当前安装的备件列表
    """
    # 验证设备存在
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    # 查询当前安装在该设备上的备件（状态为 inuse）
    instances = db.query(SparePartInstance).filter(
        SparePartInstance.installed_device_id == device_id,
        SparePartInstance.status == "inuse"
    ).all()

    return {
        "device_id": device_id,
        "device_name": device.name,
        "items": [
            {
                "instance_id": i.id,
                "serial_number": i.serial_number,
                "po_number": i.po_number,  # PO号
                "part_id": i.part_id,
                "part_number": i.part.part_number if i.part else None,
                "part_name": i.part.name if i.part else None,
                "category": i.part.category if i.part else None,
                "unit_price": float(i.unit_price) if i.unit_price else 0.0,
                "installed_at": i.installed_at.isoformat() if i.installed_at else None,
                "installed_by": i.installed_by,
                "notes": i.notes,
            }
            for i in instances
        ],
        "total": len(instances)
    }


# ============ SNMP 接口监控（阶段 A） ============

class SnmpConfig(BaseModel):
    snmp_enabled: Optional[bool] = None
    snmp_version: Optional[str] = None       # '1' | '2c'
    snmp_community: Optional[str] = None
    snmp_port: Optional[int] = None


class InterfaceUpdate(BaseModel):
    is_uplink: Optional[bool] = None
    monitored: Optional[bool] = None


def _iface_to_dict(i: DeviceInterface) -> dict:
    return {
        "id": i.id,
        "if_index": i.if_index,
        "if_name": i.if_name,
        "if_descr": i.if_descr,
        "oper_status": i.oper_status,
        "admin_status": i.admin_status,
        "speed_mbps": i.speed_mbps,
        "is_uplink": bool(i.is_uplink),
        "monitored": bool(i.monitored),
        "peer_device_id": i.peer_device_id,
        "peer_device_name": i.peer_device_name,
        "peer_ip": i.peer_ip,
        "peer_if_name": i.peer_if_name,
        "neighbor_source": i.neighbor_source,
        "neighbor_updated_at": i.neighbor_updated_at.isoformat() if i.neighbor_updated_at else None,
        "last_in_bps": i.last_in_bps,
        "last_out_bps": i.last_out_bps,
        "last_in_util": i.last_in_util,
        "last_out_util": i.last_out_util,
        "last_in_errors": i.last_in_errors,
        "last_out_errors": i.last_out_errors,
        "last_sample_at": i.last_sample_at.isoformat() if i.last_sample_at else None,
        "last_check": i.last_check.isoformat() if i.last_check else None,
    }


@router.put("/{device_id}/snmp")
async def update_device_snmp(device_id: int, cfg: SnmpConfig, db: Session = Depends(get_db)):
    """配置设备 SNMP 参数（启用/版本/团体名/端口）"""
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    if cfg.snmp_enabled is not None:
        device.snmp_enabled = cfg.snmp_enabled
    if cfg.snmp_version is not None:
        device.snmp_version = cfg.snmp_version
    if cfg.snmp_community is not None:
        device.snmp_community = cfg.snmp_community
    if cfg.snmp_port is not None:
        device.snmp_port = cfg.snmp_port
    db.commit()
    return {
        "device_id": device.id,
        "snmp_enabled": bool(device.snmp_enabled),
        "snmp_version": device.snmp_version,
        "snmp_port": device.snmp_port,
        "snmp_community_set": bool(device.snmp_community),
    }


@router.post("/{device_id}/interfaces/discover")
async def discover_device_interfaces(device_id: int, db: Session = Depends(get_db)):
    """通过 SNMP 发现设备接口（walk ifName/ifDescr），落库供标记上行口/监控"""
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    from app.services.interface_monitor import get_interface_monitor
    # discover 内部使用 asyncio.run，需在线程中执行以避免与当前事件循环冲突
    result = await asyncio.to_thread(get_interface_monitor().discover_interfaces, device_id)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail=result.get("error", "发现失败"))
    return result


@router.post("/{device_id}/interfaces/discover-neighbors")
async def discover_device_neighbors(device_id: int, db: Session = Depends(get_db)):
    """通过 CDP 自动发现邻居，回写对端关联并按层级推断上行口（第一版仅 Cisco CDP）"""
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    from app.services.interface_monitor import get_interface_monitor
    result = await asyncio.to_thread(get_interface_monitor().discover_neighbors, device_id)
    if not result.get("ok"):
        raise HTTPException(status_code=400, detail=result.get("error", "邻居发现失败"))
    return result


@router.get("/{device_id}/interfaces")
async def list_device_interfaces(device_id: int, monitored_only: bool = False, db: Session = Depends(get_db)):
    """列出设备接口"""
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    q = db.query(DeviceInterface).filter(DeviceInterface.device_id == device_id)
    if monitored_only:
        q = q.filter(DeviceInterface.monitored == True)  # noqa: E712
    interfaces = q.order_by(DeviceInterface.if_index).all()
    return {
        "device_id": device_id,
        "items": [_iface_to_dict(i) for i in interfaces],
        "total": len(interfaces),
    }


@router.put("/{device_id}/interfaces/{if_index}")
async def update_device_interface(device_id: int, if_index: int, update: InterfaceUpdate, db: Session = Depends(get_db)):
    """标记接口为上行口 / 纳入监控"""
    iface = db.query(DeviceInterface).filter(
        DeviceInterface.device_id == device_id,
        DeviceInterface.if_index == if_index,
    ).first()
    if not iface:
        raise HTTPException(status_code=404, detail="接口不存在，请先发现接口")

    if update.is_uplink is not None:
        iface.is_uplink = update.is_uplink
    if update.monitored is not None:
        iface.monitored = update.monitored
    db.commit()
    return _iface_to_dict(iface)


@router.get("/{device_id}/interfaces/{if_index}/traffic")
async def get_interface_traffic(device_id: int, if_index: int, limit: int = 60, db: Session = Depends(get_db)):
    """获取接口最近流量样本（默认最近 60 个点，倒序时间）"""
    iface = db.query(DeviceInterface).filter(
        DeviceInterface.device_id == device_id,
        DeviceInterface.if_index == if_index,
    ).first()
    if not iface:
        raise HTTPException(status_code=404, detail="接口不存在")

    limit = max(1, min(limit, 500))
    samples = db.query(InterfaceTrafficSample).filter(
        InterfaceTrafficSample.interface_id == iface.id,
    ).order_by(InterfaceTrafficSample.ts.desc()).limit(limit).all()

    return {
        "device_id": device_id,
        "if_index": if_index,
        "if_name": iface.if_name,
        "interface": _iface_to_dict(iface),
        "samples": [
            {
                "ts": s.ts.isoformat() if s.ts else None,
                "in_bps": s.in_bps,
                "out_bps": s.out_bps,
                "in_util": s.in_util,
                "out_util": s.out_util,
                "in_errors": s.in_errors,
                "out_errors": s.out_errors,
                "oper_status": s.oper_status,
            }
            for s in reversed(samples)
        ],
    }


@router.post("/monitor/discover-neighbors-all")
async def discover_neighbors_all(db: Session = Depends(get_db)):
    """对所有启用 SNMP 的设备批量执行 CDP/LLDP 邻居发现"""
    devices = db.query(Device).filter(
        Device.snmp_enabled == True,  # noqa: E712
        Device.ip.isnot(None),
        Device.snmp_community.isnot(None),
    ).all()
    if not devices:
        return {"ok": True, "devices": 0, "results": []}

    from app.services.interface_monitor import get_interface_monitor
    monitor = get_interface_monitor()

    results = []
    total_found = total_matched = total_uplinks = 0
    for dev in devices:
        res = await asyncio.to_thread(monitor.discover_neighbors, dev.id)
        results.append({
            "device_id": dev.id,
            "device_name": dev.name,
            "ok": res.get("ok", False),
            "found": res.get("found", 0),
            "matched": res.get("matched", 0),
            "uplinks_marked": res.get("uplinks_marked", 0),
            "error": res.get("error"),
        })
        if res.get("ok"):
            total_found += res.get("found", 0)
            total_matched += res.get("matched", 0)
            total_uplinks += res.get("uplinks_marked", 0)

    return {
        "ok": True,
        "devices": len(devices),
        "total_found": total_found,
        "total_matched": total_matched,
        "total_uplinks_marked": total_uplinks,
        "results": results,
    }


@router.get("/monitor/neighbor-links")
async def list_neighbor_links(db: Session = Depends(get_db)):
    """返回 CDP/LLDP 发现的邻居链路（供大屏自动绘制拓扑线）"""
    rows = db.query(DeviceInterface).filter(
        DeviceInterface.neighbor_source.isnot(None),
    ).all()
    links = []
    for i in rows:
        links.append({
            "device_id": i.device_id,
            "if_index": i.if_index,
            "if_name": i.if_name,
            "oper_status": i.oper_status,
            "is_uplink": bool(i.is_uplink),
            "peer_device_id": i.peer_device_id,
            "peer_device_name": i.peer_device_name,
            "peer_ip": i.peer_ip,
            "peer_if_name": i.peer_if_name,
            "source": i.neighbor_source,
            "updated_at": i.neighbor_updated_at.isoformat() if i.neighbor_updated_at else None,
        })
    return {"count": len(links), "links": links}
