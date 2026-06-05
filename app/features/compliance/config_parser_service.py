"""
配置文件解析服务

支持格式：
- .txt - 文本文件
- .cfg - Cisco 配置文件
- .xlsx - Excel 批量设备配置
- .log - 日志文件（可作为配置文件）

解析后的配置文本供 AI 审核使用。
"""
import re
import io
from typing import List, Dict, Optional
from pathlib import Path
from loguru import logger
from datetime import datetime

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl 未安装，Excel 解析功能不可用")


class ConfigParserService:
    """配置文件解析服务"""

    SUPPORTED_EXTENSIONS = ['.txt', '.cfg', '.log', '.xlsx', '.xls', '.conf']

    def parse_file(self, file_content: bytes, filename: str) -> Dict:
        """
        解析上传的配置文件

        Args:
            file_content: 文件二进制内容
            filename: 文件名（用于判断格式）

        Returns:
            解析结果，包含配置文本和元数据
        """
        ext = Path(filename).suffix.lower()

        if ext not in self.SUPPORTED_EXTENSIONS:
            return {
                "success": False,
                "error": f"不支持的文件格式: {ext}，支持格式: {', '.join(self.SUPPORTED_EXTENSIONS)}"
            }

        try:
            if ext in ['.xlsx', '.xls']:
                return self._parse_excel(file_content)
            else:
                return self._parse_text(file_content, ext)

        except Exception as e:
            logger.error(f"文件解析失败: {e}")
            return {
                "success": False,
                "error": f"文件解析失败: {str(e)}"
            }

    def _parse_text(self, file_content: bytes, ext: str) -> Dict:
        """
        解析文本配置文件（txt, cfg, log, conf）

        Args:
            file_content: 文件内容
            ext: 文件扩展名

        Returns:
            解析结果
        """
        # 尝试多种编码
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
        text_content = None
        used_encoding = None

        for encoding in encodings:
            try:
                text_content = file_content.decode(encoding)
                used_encoding = encoding
                break
            except UnicodeDecodeError:
                continue

        if text_content is None:
            return {
                "success": False,
                "error": "无法解码文件内容，请确保文件是有效的文本格式"
            }

        # 清理内容
        lines = text_content.strip().split('\n')

        # 统计信息
        total_lines = len(lines)
        config_lines = sum(1 for line in lines if line.strip() and not line.startswith('!') and not line.startswith('#'))

        # 识别设备类型（Cisco 配置特征）
        device_type = self._detect_device_type(text_content)

        # 提取关键信息
        hostname = self._extract_hostname(text_content)

        result = {
            "success": True,
            "config_text": text_content,
            "device_type": device_type,
            "hostname": hostname,
            "total_lines": total_lines,
            "config_lines": config_lines,
            "encoding": used_encoding,
            "file_format": ext,
            "parse_time": datetime.utcnow().isoformat()
        }

        logger.info(f"文本配置解析成功: {total_lines} 行, {config_lines} 配置行, 设备类型: {device_type}")

        return result

    def _parse_excel(self, file_content: bytes) -> Dict:
        """
        解析 Excel 文件中的设备配置

        Excel 格式可以是：
        1. 单设备配置 - 一列或多列包含配置内容
        2. 多设备配置 - 每行一个设备，包含设备名称和配置内容

        Args:
            file_content: Excel 文件内容

        Returns:
            解析结果
        """
        if not EXCEL_AVAILABLE:
            return {
                "success": False,
                "error": "Excel 解析功能不可用，请安装 openpyxl"
            }

        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active

        # 读取表头
        headers = [cell.value for cell in ws[1] if cell.value]
        logger.debug(f"Excel 表头: {headers}")

        # 判断 Excel 格式
        if 'config' in headers or '配置' in headers or '配置内容' in headers:
            # 多设备配置格式
            return self._parse_multi_device_excel(ws, headers)
        else:
            # 单设备配置格式（所有内容作为配置）
            return self._parse_single_device_excel(ws)

    def _parse_multi_device_excel(self, ws, headers: List[str]) -> Dict:
        """
        解析多设备配置 Excel

        期望格式：
        | 设备名称 | IP | 配置内容 | 备注 |
        | SW-01 | 1.1.1.1 | hostname SW-01... | ... |

        Args:
            ws: Excel worksheet
            headers: 表头列表

        Returns:
            解析结果，包含多设备配置
        """
        # 找到配置列
        config_col_idx = None
        device_col_idx = None
        ip_col_idx = None

        for i, header in enumerate(headers):
            header_lower = str(header).lower()
            if 'config' in header_lower or '配置' in header_lower or '配置内容' in header_lower:
                config_col_idx = i + 1  # Excel 列索引从 1 开始
            elif 'device' in header_lower or '设备' in header_lower or '名称' in header_lower:
                device_col_idx = i + 1
            elif 'ip' in header_lower:
                ip_col_idx = i + 1

        if config_col_idx is None:
            return {
                "success": False,
                "error": "Excel 中未找到配置内容列，请确保表头包含 'config' 或 '配置内容'"
            }

        devices = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if not any(row):
                continue

            device_name = row[device_col_idx - 1] if device_col_idx else f"Device-{row_idx}"
            ip_address = row[ip_col_idx - 1] if ip_col_idx else None
            config_text = row[config_col_idx - 1]

            if config_text:
                devices.append({
                    "device_name": str(device_name) if device_name else f"Device-{row_idx}",
                    "ip": str(ip_address) if ip_address else None,
                    "config_text": str(config_text),
                    "row": row_idx
                })

        if not devices:
            return {
                "success": False,
                "error": "Excel 中未找到有效的配置数据"
            }

        # 合并所有配置文本（用于整体审核）
        combined_config = "\n\n".join([
            f"! Device: {d['device_name']}\n{d['config_text']}"
            for d in devices
        ])

        result = {
            "success": True,
            "config_text": combined_config,
            "devices": devices,
            "device_count": len(devices),
            "format": "multi_device",
            "parse_time": datetime.utcnow().isoformat()
        }

        logger.info(f"Excel 多设备配置解析成功: {len(devices)} 台设备")

        return result

    def _parse_single_device_excel(self, ws) -> Dict:
        """
        解析单设备配置 Excel

        将所有单元格内容合并为配置文本

        Args:
            ws: Excel worksheet

        Returns:
            解析结果
        """
        lines = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and str(cell).strip():
                    lines.append(str(cell).strip())

        config_text = '\n'.join(lines)

        result = {
            "success": True,
            "config_text": config_text,
            "total_lines": len(lines),
            "format": "single_device",
            "parse_time": datetime.utcnow().isoformat()
        }

        logger.info(f"Excel 单设备配置解析成功: {len(lines)} 行")

        return result

    def _detect_device_type(self, config_text: str) -> str:
        """
        检测设备类型/厂商

        Args:
            config_text: 配置文本

        Returns:
            设备类型 (cisco_ios, juniper_junos, huawei, unknown)
        """
        # Cisco IOS 特征
        cisco_patterns = [
            r'^hostname\s+',
            r'^interface\s+\w+',
            r'^ip\s+address\s+',
            r'^router\s+ospf',
            r'^line\s+console',
            r'^line\s+vty',
            r'^enable\s+secret',
            r'^!\s*$',
        ]

        # Juniper 特征
        juniper_patterns = [
            r'system\s+{',
            r'interfaces\s+{',
            r'ge-\d+/\d+/\d+',
            r'family\s+inet',
            r'junos',
        ]

        # Huawei 特征
        huawei_patterns = [
            r'sysname\s+',
            r'interface\s+GigabitEthernet',
            r'huawei',
        ]

        config_lower = config_text.lower()

        # 统计特征匹配
        cisco_count = sum(1 for p in cisco_patterns if re.search(p, config_text, re.MULTILINE))
        juniper_count = sum(1 for p in juniper_patterns if re.search(p, config_text, re.IGNORECASE))
        huawei_count = sum(1 for p in huawei_patterns if re.search(p, config_text, re.IGNORECASE))

        if cisco_count >= 3:
            return "cisco_ios"
        elif juniper_count >= 2:
            return "juniper_junos"
        elif huawei_count >= 2:
            return "huawei"
        else:
            return "unknown"

    def _extract_hostname(self, config_text: str) -> Optional[str]:
        """
        提取设备 hostname

        Args:
            config_text: 配置文本

        Returns:
            hostname 或 None
        """
        # Cisco: hostname SW-01
        match = re.search(r'^hostname\s+(\S+)', config_text, re.MULTILINE)
        if match:
            return match.group(1)

        # Huawei: sysname SW-01
        match = re.search(r'^sysname\s+(\S+)', config_text, re.MULTILINE)
        if match:
            return match.group(1)

        return None

    def split_configs(self, combined_config: str) -> List[Dict]:
        """
        将合并的配置拆分为单个设备配置

        Args:
            combined_config: 合并的配置文本（包含多个设备）

        Returns:
            单个设备配置列表
        """
        # 按 "! Device:" 分隔符拆分
        pattern = r'! Device:\s*(\S+)'
        parts = re.split(pattern, combined_config)

        devices = []
        # parts 格式: ['', 'SW-01', 'config1', 'SW-02', 'config2', ...]
        for i in range(1, len(parts), 2):
            if i + 1 < len(parts):
                device_name = parts[i]
                config_text = parts[i + 1].strip()
                devices.append({
                    "device_name": device_name,
                    "config_text": config_text
                })

        return devices

    def get_config_summary(self, config_text: str) -> Dict:
        """
        获取配置摘要信息

        Args:
            config_text: 配置文本

        Returns:
            配置摘要
        """
        lines = config_text.strip().split('\n')

        # 统计各类配置项
        interfaces = []
        for line in lines:
            match = re.match(r'^interface\s+(\S+)', line)
            if match:
                interfaces.append(match.group(1))

        # VLAN 统计
        vlans = []
        for line in lines:
            match = re.match(r'^vlan\s+(\d+)', line)
            if match:
                vlans.append(match.group(1))

        return {
            "total_lines": len(lines),
            "interfaces": interfaces,
            "interface_count": len(interfaces),
            "vlans": vlans,
            "vlan_count": len(vlans),
            "device_type": self._detect_device_type(config_text),
            "hostname": self._extract_hostname(config_text)
        }